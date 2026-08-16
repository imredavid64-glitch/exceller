"""
Excel Parser - File reading and analysis
Handles Excel file operations and data extraction
"""

import os
from pathlib import Path
from typing import Dict, List, Any, Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None


class ExcelParser:
    """Excel file parser and analyzer"""
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls', '.csv']
    
    def get_file_stats(self, file_path: str) -> Dict[str, Any]:
        """Get basic statistics about an Excel file"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        suffix = file_path.suffix.lower()
        
        if suffix == '.csv':
            return self._get_csv_stats(file_path)
        elif suffix in ['.xlsx', '.xls']:
            return self._get_excel_stats(file_path)
        else:
            raise ValueError(f"Unsupported file format: {suffix}")
    
    def _get_excel_stats(self, file_path: Path) -> Dict[str, Any]:
        """Get statistics for Excel files"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            total_rows = 0
            total_cols = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                total_rows += ws.max_row or 0
                total_cols = max(total_cols, ws.max_column or 0)
            
            wb.close()
            
            return {
                'sheets': len(wb.sheetnames),
                'rows': total_rows,
                'columns': total_cols,
                'sheet_names': wb.sheetnames,
                'file_size': file_path.stat().st_size
            }
            
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def _get_csv_stats(self, file_path: Path) -> Dict[str, Any]:
        """Get statistics for CSV files"""
        try:
            if pd:
                df = pd.read_csv(file_path)
                return {
                    'sheets': 1,
                    'rows': len(df),
                    'columns': len(df.columns),
                    'sheet_names': ['Sheet1'],
                    'file_size': file_path.stat().st_size,
                    'column_names': list(df.columns)
                }
            else:
                # Fallback without pandas
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    lines = f.readlines()
                    rows = len(lines) - 1  # Subtract header
                    cols = len(lines[0].split(',')) if lines else 0
                    return {
                        'sheets': 1,
                        'rows': rows,
                        'columns': cols,
                        'sheet_names': ['Sheet1'],
                        'file_size': file_path.stat().st_size
                    }
                    
        except Exception as e:
            raise Exception(f"Error reading CSV file: {str(e)}")
    
    def get_file_context(self, file_path: str, max_rows: int = 20) -> str:
        """Get file content as context for AI"""
        file_path = Path(file_path)
        suffix = file_path.suffix.lower()
        
        if suffix == '.csv':
            return self._get_csv_context(file_path, max_rows)
        elif suffix in ['.xlsx', '.xls']:
            return self._get_excel_context(file_path, max_rows)
        else:
            return "Unsupported file format"
    
    def _get_excel_context(self, file_path: Path, max_rows: int) -> str:
        """Get Excel file context"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            context_parts = []
            
            for sheet_name in wb.sheetnames[:3]:  # Limit to first 3 sheets
                ws = wb[sheet_name]
                
                # Get headers (first row)
                headers = []
                for col in range(1, min(ws.max_column or 0, 20) + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value else f"Column{col}")
                
                # Get sample data
                sample_data = []
                for row in range(2, min(max_rows + 1, (ws.max_row or 0) + 1)):
                    row_data = []
                    for col in range(1, min(ws.max_column or 0, 20) + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(str(cell_value) if cell_value else "")
                    sample_data.append(row_data)
                
                # Build context
                sheet_context = f"\n=== Sheet: {sheet_name} ===\n"
                sheet_context += f"Headers: {' | '.join(headers)}\n"
                sheet_context += f"Total Rows: {ws.max_row}\n"
                sheet_context += "Sample Data:\n"
                
                for i, row in enumerate(sample_data, 1):
                    sheet_context += f"Row {i}: {' | '.join(row)}\n"
                
                context_parts.append(sheet_context)
            
            wb.close()
            
            return "\n".join(context_parts)
            
        except Exception as e:
            return f"Error reading file: {str(e)}"
    
    def _get_csv_context(self, file_path: Path, max_rows: int) -> str:
        """Get CSV file context"""
        try:
            if pd:
                df = pd.read_csv(file_path, nrows=max_rows)
                
                context = f"\n=== CSV File ===\n"
                context += f"Headers: {' | '.join(df.columns.tolist())}\n"
                context += f"Total Rows: {len(df)}\n"
                context += "Sample Data:\n"
                
                for idx, row in df.head(max_rows).iterrows():
                    context += f"Row {idx + 1}: {' | '.join(row.astype(str).tolist())}\n"
                
                return context
            else:
                # Fallback without pandas
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    lines = f.readlines()
                    
                    context = f"\n=== CSV File ===\n"
                    context += f"Headers: {lines[0].strip()}\n"
                    context += f"Total Rows: {len(lines) - 1}\n"
                    context += "Sample Data:\n"
                    
                    for i, line in enumerate(lines[1:max_rows + 1], 1):
                        context += f"Row {i}: {line.strip()}\n"
                    
                    return context
                    
        except Exception as e:
            return f"Error reading CSV: {str(e)}"
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """Get list of sheet names in Excel file"""
        file_path = Path(file_path)
        
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                return sheet_names
            except Exception as e:
                return []
        else:
            return ['Sheet1']
    
    def read_sheet_data(self, file_path: str, sheet_name: str = None, 
                        max_rows: int = None) -> Dict[str, Any]:
        """Read data from a specific sheet"""
        file_path = Path(file_path)
        
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                
                # Read data
                data = []
                headers = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx == 0:
                        headers = [str(cell) if cell else f"Col{i}" for i, cell in enumerate(row, 1)]
                    else:
                        if max_rows and row_idx >= max_rows + 1:
                            break
                        data.append([str(cell) if cell else "" for cell in row])
                
                wb.close()
                
                return {
                    'headers': headers,
                    'data': data,
                    'total_rows': ws.max_row,
                    'sheet_name': ws.title
                }
                
            except Exception as e:
                return {'error': str(e)}
        else:
            return {'error': 'Unsupported file format'}
    
    def find_issues(self, file_path: str) -> List[Dict[str, Any]]:
        """Find potential issues in the data"""
        issues = []
        
        try:
            context = self.get_file_context(file_path)
            stats = self.get_file_stats(file_path)
            
            # Basic issue detection
            if stats['rows'] == 0:
                issues.append({
                    'type': 'empty',
                    'severity': 'high',
                    'message': 'File appears to be empty'
                })
            
            if stats['columns'] == 0:
                issues.append({
                    'type': 'no_columns',
                    'severity': 'high',
                    'message': 'No columns detected'
                })
            
            return issues
            
        except Exception as e:
            return [{'type': 'error', 'severity': 'high', 'message': str(e)}]
